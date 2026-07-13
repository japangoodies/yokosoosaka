#!/usr/bin/env python3
"""
rag_indexer.py — Index Desktop files into ChromaDB for RAG
Usage: python rag_indexer.py [--rebuild] [--dry-run]
"""

import json
import argparse
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from langchain_text_splitters import RecursiveCharacterTextSplitter
import chromadb
from chromadb import Documents, EmbeddingFunction, Embeddings
from sentence_transformers import SentenceTransformer
import docx
import fitz
import openpyxl
import pandas as pd


MODEL_NAME = "paraphrase-MiniLM-L3-v2"
CHUNK_SIZE = 2000
CHUNK_OVERLAP = 200
COLLECTION_NAME = "desktop_files"
BATCH_SIZE = 500

RAG_ROOT = Path("E:/opencode") / "rag_system"
CHROMA_DIR = RAG_ROOT / "chroma_db"
STATE_FILE = RAG_ROOT / "index_state.json"

INDEX_FOLDERS = [
    "HO MODULE",
    "PM",
    "Brownies Unlimited",
    "FIT Training",
]

# Path substrings to EXCLUDE (case-insensitive)
EXCLUDE_PATTERNS = [
    "node_modules",
    "\\.git",
    "__pycache__",
    "opencode",
    ".opencode",
    "DATA STRUCTURE",  # VQPBOS table schemas
    "mall93",
    "con93",
    "deskbak",
]

SUPPORTED_EXTENSIONS = {
    ".docx", ".pdf", ".xlsx", ".xls", ".csv",
    ".txt", ".py", ".sql", ".json", ".md",
    ".yaml", ".yml", ".toml", ".cfg", ".ini", ".conf",
    ".log", ".xml", ".rtf",
}


class SentenceTransformerEmbeddingFunction(EmbeddingFunction):
    def __init__(self, model_name: str = MODEL_NAME):
        self.model = SentenceTransformer(model_name)

    def __call__(self, input: Documents) -> Embeddings:
        return self.model.encode(list(input)).tolist()


class BatchEmbedder:
    def __init__(self, model_name: str = MODEL_NAME, device: str = "cpu"):
        self.model = SentenceTransformer(model_name, device=device)

    def embed(self, texts: List[str], batch_size: int = BATCH_SIZE) -> List[List[float]]:
        return self.model.encode(texts, batch_size=batch_size, show_progress_bar=True).tolist()


def parse_docx(filepath: Path) -> str:
    doc = docx.Document(str(filepath))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def parse_pdf(filepath: Path) -> str:
    text_parts = []
    with fitz.open(str(filepath)) as doc:
        for page in doc:
            text_parts.append(page.get_text())
    return "\n".join(text_parts)


def parse_xlsx(filepath: Path) -> str:
    try:
        wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
    except openpyxl.utils.exceptions.InvalidFileException:
        return parse_xls(filepath)
    text_parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            rows.append(" | ".join(cells))
        text_parts.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows))
    wb.close()
    return "\n".join(text_parts)


def parse_xls(filepath: Path) -> str:
    import xlrd
    wb = xlrd.open_workbook(str(filepath))
    text_parts = []
    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        rows = []
        for row_idx in range(ws.nrows):
            cells = [str(ws.cell_value(row_idx, c)) for c in range(ws.ncols)]
            rows.append(" | ".join(cells))
        text_parts.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows))
    return "\n".join(text_parts)


def parse_csv(filepath: Path) -> str:
    df = pd.read_csv(filepath)
    return df.to_string()


def parse_text(filepath: Path) -> str:
    try:
        return filepath.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return filepath.read_text(encoding="latin-1")


PARSERS = {
    ".docx": parse_docx, ".doc": parse_docx,
    ".pdf": parse_pdf,
    ".xlsx": parse_xlsx, ".xls": parse_xlsx,
    ".csv": parse_csv,
    ".txt": parse_text, ".py": parse_text, ".sql": parse_text,
    ".json": parse_text, ".md": parse_text,
    ".yaml": parse_text, ".yml": parse_text,
    ".toml": parse_text, ".cfg": parse_text,
    ".ini": parse_text, ".conf": parse_text,
    ".log": parse_text, ".xml": parse_text,
    ".html": parse_text, ".css": parse_text, ".js": parse_text,
    ".rtf": parse_text,
}


def parse_file(filepath: Path) -> Optional[str]:
    ext = filepath.suffix.lower()
    parser = PARSERS.get(ext)
    if parser is None:
        return None
    try:
        return parser(filepath)
    except Exception as e:
        print(f"  [!] Error parsing {filepath.name}: {e}")
        return None


def get_file_signature(filepath: Path) -> str:
    stat = filepath.stat()
    return f"{filepath}::{stat.st_mtime}::{stat.st_size}"


def discover_files(root_dirs: List[Path]) -> List[Path]:
    files = []
    for root in root_dirs:
        if not root.exists():
            print(f"  [!] Directory not found: {root}")
            continue
        for entry in root.rglob("*"):
            if not entry.is_file():
                continue
            if entry.name.startswith("~$"):
                continue
            if entry.suffix.lower() not in SUPPORTED_EXTENSIONS:
                continue

            rel = str(entry.relative_to(Path.home())).lower()
            excluded = False
            for pat in EXCLUDE_PATTERNS:
                if pat.lower() in rel:
                    excluded = True
                    break
            if excluded:
                continue

            files.append(entry)
    return files


def index_files(files: List[Path], rebuild: bool = False):
    RAG_ROOT.mkdir(parents=True, exist_ok=True)
    CHROMA_DIR.mkdir(parents=True, exist_ok=True)

    state = {}
    if not rebuild and STATE_FILE.exists():
        state = json.loads(STATE_FILE.read_text())

    new_state = {}
    to_index = []

    for filepath in files:
        sig = get_file_signature(filepath)
        new_state[str(filepath)] = sig
        if sig != state.get(str(filepath)):
            to_index.append(filepath)

    if not to_index:
        print("  All files are up to date. Nothing to index.")
        return

    print(f"  Found {len(to_index)} file(s) to index (out of {len(files)} total)")

    text_splitter = RecursiveCharacterTextSplitter(
        chunk_size=CHUNK_SIZE,
        chunk_overlap=CHUNK_OVERLAP,
        length_function=len,
        separators=["\n\n", "\n", ".", " ", ""],
    )

    all_texts = []
    all_ids = []
    all_metadatas = []

    print("  Phase 1: Parsing files...")
    for i, filepath in enumerate(to_index):
        rel_path = filepath.relative_to(Path.home())
        print(f"    [{i+1}/{len(to_index)}] {rel_path}")

        text = parse_file(filepath)
        if text is None or not text.strip():
            continue
        if len(text.strip()) < 50:
            continue

        stat = filepath.stat()
        chunks = text_splitter.split_text(text)

        for j, chunk in enumerate(chunks):
            doc_id = f"{filepath}::chunk::{j}"
            all_texts.append(chunk)
            all_ids.append(doc_id)
            all_metadatas.append({
                "source": str(filepath),
                "filename": filepath.name,
                "extension": filepath.suffix.lower(),
                "folder": str(filepath.parent),
                "chunk_index": j,
                "total_chunks": len(chunks),
                "size": stat.st_size,
                "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
            })

    print(f"\n  Phase 2: Computing embeddings for {len(all_texts)} chunks...")
    embedder = BatchEmbedder()
    all_embeddings = embedder.embed(all_texts)

    print("  Phase 3: Storing in ChromaDB...")
    client = chromadb.PersistentClient(str(CHROMA_DIR))

    if rebuild:
        try:
            client.delete_collection(COLLECTION_NAME)
            print("    Rebuilt: deleted existing collection")
        except ValueError:
            pass

    collection = client.get_or_create_collection(
        name=COLLECTION_NAME,
        embedding_function=SentenceTransformerEmbeddingFunction(),
    )

    for i in range(0, len(all_texts), BATCH_SIZE):
        end = i + BATCH_SIZE
        print(f"    Adding batch {i//BATCH_SIZE + 1}/{(len(all_texts)-1)//BATCH_SIZE + 1}...")
        collection.add(
            documents=all_texts[i:end],
            embeddings=all_embeddings[i:end],
            ids=all_ids[i:end],
            metadatas=all_metadatas[i:end],
        )

    STATE_FILE.write_text(json.dumps(new_state, indent=2))

    count = collection.count()
    print(f"\n  Done! Collection has {count} chunks.")


def main():
    parser = argparse.ArgumentParser(description="Index Desktop files into ChromaDB for RAG")
    parser.add_argument("--rebuild", action="store_true", help="Delete and rebuild the entire index")
    parser.add_argument("--dry-run", action="store_true", help="Show which files would be indexed")
    args = parser.parse_args()

    desktop = Path.home() / "Desktop"
    root_dirs = [desktop / folder for folder in INDEX_FOLDERS]
    root_dirs.append(desktop)

    print("Discovering files...")
    files = discover_files(root_dirs)
    print(f"  Found {len(files)} supported files in {len(INDEX_FOLDERS)+1} folder(s)")

    if args.dry_run:
        print("\nFiles to index:")
        for f in sorted(files):
            print(f"  - {f.relative_to(Path.home())}")
        return

    print("\nIndexing files...")
    index_files(files, rebuild=args.rebuild)


if __name__ == "__main__":
    main()
